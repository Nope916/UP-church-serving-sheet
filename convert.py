from openpyxl import load_workbook
from html import escape
from pathlib import Path
from tkinter import Tk, filedialog
from datetime import datetime, date

root = Tk()
root.withdraw()

xlsx_path_str = filedialog.askopenfilename(
    title="請選擇 / 拖曳 服事表 Excel",
    filetypes=[("Excel files", "*.xlsx")]
)
root.destroy()

if not xlsx_path_str:
    raise SystemExit("沒有選擇 Excel")

xlsx_path = Path(xlsx_path_str)
out_path = xlsx_path.with_name(xlsx_path.stem + "_table_snippet.html")

wb = load_workbook(xlsx_path, data_only=True)
ws = wb.active

BR_TOKEN = "__BR__"

def is_blank(v):
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False

def cell_text(r, c):
    v = ws.cell(r, c).value
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")

    s = str(v).strip()
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", BR_TOKEN)  # 先用 token 佔位，避免被 escape
    return s

merged_map = {}
for mr in ws.merged_cells.ranges:
    merged_map[(mr.min_row, mr.min_col)] = (
        mr.max_row - mr.min_row + 1,
        mr.max_col - mr.min_col + 1
    )
    for r in range(mr.min_row, mr.max_row + 1):
        for c in range(mr.min_col, mr.max_col + 1):
            if (r, c) != (mr.min_row, mr.min_col):
                merged_map[(r, c)] = None

last_row = 0
last_col = 0
for row in ws.iter_rows():
    for cell in row:
        if not is_blank(cell.value):
            last_row = max(last_row, cell.row)
            last_col = max(last_col, cell.column)

max_row = last_row
max_col = last_col

def render_cell(tag, r, c):
    mi = merged_map.get((r, c), "not_merged")
    if mi is None:
        return None

    rs = cs = 1
    if mi != "not_merged":
        rs, cs = mi

    txt = cell_text(r, c)
    txt = escape(txt, quote=False)          # 先 escape
    txt = txt.replace(BR_TOKEN, "<br>")     # 再把 token 換回 <br>

    attrs = []
    if rs > 1:
        attrs.append(f'rowspan="{rs}"')
    if cs > 1:
        attrs.append(f'colspan="{cs}"')

    if attrs:
        return f"<{tag} {' '.join(attrs)}>{txt}</{tag}>"
    return f"<{tag}>{txt}</{tag}>"

def build_row(r, tag):
    cells = []
    for c in range(1, max_col + 1):
        out = render_cell(tag, r, c)
        if out is None:
            continue
        cells.append(out)
    return cells

thead_cells = build_row(1, "th")
thead_html = "<thead>\n  <tr>\n    " + "\n    ".join(thead_cells) + "\n  </tr>\n</thead>"

tbody_rows = []
for r in range(2, max_row + 1):
    tds = build_row(r, "td")
    if not tds:
        continue
    tbody_rows.append("  <tr>\n    " + "\n    ".join(tds) + "\n  </tr>")
tbody_html = "<tbody>\n" + "\n".join(tbody_rows) + "\n</tbody>"

snippet = f"""<div class="table-wrap">
  <table id="roster">
{thead_html}
{tbody_html}
  </table>
</div>
"""

out_path.write_text(snippet, encoding="utf-8")
print("Wrote:", out_path)
